"""Extract structured crime statistics from ACT Policing monthly XLSX export."""

from __future__ import annotations

import json
import re
import sys
from pathlib import Path

from openpyxl import load_workbook

MONTHS = {
    "JAN": 1,
    "FEB": 2,
    "MAR": 3,
    "APR": 4,
    "MAY": 5,
    "JUN": 6,
    "JUL": 7,
    "AUG": 8,
    "SEP": 9,
    "OCT": 10,
    "NOV": 11,
    "DEC": 12,
}

TABLE_OFFENCE_RE = re.compile(
    r"^Table\s+\d+:\s*(.+?)\s*-\s*Offences and other activities\s*$",
    re.I,
)
TABLE_ANY_RE = re.compile(r"^Table\s+(\d+):\s*(.+)\s*$", re.I)


def parse_period_label(label: str) -> tuple[int, int] | None:
    """MONYY -> (year, month) e.g. MAR26 -> (2026, 3)."""
    if not label or not isinstance(label, str):
        return None
    label = label.strip().upper()
    m = re.match(r"^(JAN|FEB|MAR|APR|MAY|JUN|JUL|AUG|SEP|OCT|NOV|DEC)(\d{2})$", label)
    if not m:
        return None
    mon = MONTHS[m.group(1)]
    yy = int(m.group(2))
    year = 2000 + yy if yy <= 50 else 1900 + yy
    return year, mon


def period_sort_key(label: str) -> tuple[int, int]:
    p = parse_period_label(label)
    return p if p else (0, 0)


def district_from_offence_title(title: str) -> str | None:
    m = TABLE_OFFENCE_RE.match(title.strip())
    return m.group(1).strip() if m else None


def _is_offence_header_row(r: tuple) -> bool:
    a = r[0] if r else None
    b = r[1] if r and len(r) > 1 else None
    if a in ("Offence", "Offences") and b == "Date reported":
        return True
    if isinstance(a, str) and a.strip() == "Date reported":
        return True
    if (a is None or (isinstance(a, str) and not str(a).strip())) and b == "Date reported":
        return True
    return False


def extract_offence_tables(ws) -> list[dict]:
    """Return list of {district, periods: [...], series: {offence: {period: int}}}."""
    rows = list(ws.iter_rows(values_only=True))
    tables: list[dict] = []
    i = 0
    n = len(rows)
    while i < n:
        row = rows[i]
        c0 = row[0] if row else None
        if isinstance(c0, str) and c0.startswith("Table ") and "Offences and other activities" in c0:
            title = c0
            district = district_from_offence_title(title)
            if not district:
                i += 1
                continue
            j = i + 1
            periods: list[str] | None = None
            series: dict[str, dict[str, int]] = {}
            while j < n:
                r = rows[j]
                first = r[0] if r else None
                if isinstance(first, str) and first.startswith("Table ") and j > i + 3:
                    break
                if _is_offence_header_row(r):
                    header = rows[j + 1] if j + 1 < n else None
                    if header is not None and (
                        header[0] is None or str(header[0]).strip() == ""
                    ):
                        periods = []
                        for cell in header[1:]:
                            if cell is None:
                                break
                            periods.append(str(cell).strip())
                    j += 2
                    continue
                if (
                    periods
                    and first
                    and isinstance(first, str)
                    and first not in ("Offence", "Offences")
                    and not (isinstance(first, str) and first.strip() == "Date reported")
                ):
                    name = first.strip()
                    data: dict[str, int] = {}
                    for k, p in enumerate(periods):
                        val = r[k + 1] if k + 1 < len(r) else None
                        if val is None:
                            continue
                        try:
                            data[p] = int(val)
                        except (TypeError, ValueError):
                            data[p] = 0
                    series[name] = data
                j += 1
            if periods:
                tables.append(
                    {
                        "district": district,
                        "title": title,
                        "periods": periods,
                        "series": series,
                    }
                )
            i = j
            continue
        i += 1
    return tables


def extract_simple_table(ws, header_row_idx: int) -> tuple[str | None, list[str], dict[str, dict[str, int]]]:
    """Returns (table_title from col A if present), periods, series."""
    rows = list(ws.iter_rows(values_only=True))
    title: str | None = None
    if header_row_idx > 0:
        t0 = rows[0][0] if rows[0] else None
        if isinstance(t0, str) and t0.strip().startswith("Table "):
            title = t0.strip()
    header = rows[header_row_idx]
    periods: list[str] = []
    for cell in header[1:]:
        if cell is None or (isinstance(cell, str) and not cell.strip()):
            break
        periods.append(str(cell).strip())
    series: dict[str, dict[str, int]] = {}
    for r in rows[header_row_idx + 1 :]:
        label = r[0]
        if label is None:
            continue
        label = str(label).strip()
        if not label or label == " ":
            continue
        data: dict[str, int] = {}
        for k, p in enumerate(periods):
            val = r[k + 1] if k + 1 < len(r) else None
            try:
                data[p] = int(val) if val is not None else 0
            except (TypeError, ValueError):
                data[p] = 0
        series[label] = data
    return title, periods, series


def merge_periods(*period_lists: list[str]) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for lst in period_lists:
        for p in lst:
            if p not in seen:
                seen.add(p)
                out.append(p)
    return sorted(out, key=period_sort_key)


def table_number_from_title(title: str | None) -> int | None:
    if not title:
        return None
    m = TABLE_ANY_RE.match(title.strip())
    return int(m.group(1)) if m else None


# --- Quarterly suburb (community) workbook (partial timeline vs monthly MARYY) ---

COMMUNITY_SHEET_TO_DISTRICT: dict[str, str] = {
    "Belconnen": "Belconnen",
    "Gungahlin": "Gungahlin",
    "Inner North": "Inner North",
    "Inner South": "Inner South",
    "Weston Creek": "Weston",
    "Molonglo District": "Molonglo District",
    "Woden": "Woden",
    "Tuggeranong": "Tuggeranong",
    "Other": "Other Areas",
}

QUARTER_HDR_RE = re.compile(r"^\d{4}\s+Q[1-4]\b", re.I)


def _is_quarter_header_row(row: tuple) -> bool:
    a = row[0] if row else None
    b = row[1] if row and len(row) > 1 else None
    if a is not None and str(a).strip():
        return False
    if not b:
        return False
    return bool(QUARTER_HDR_RE.match(str(b).strip()))


def _offence_block_follows(rows: list[tuple], i: int) -> bool:
    if i + 1 >= len(rows):
        return False
    return _is_quarter_header_row(rows[i + 1])


def _parse_suburb_value_row(
    row: tuple, num_periods: int
) -> tuple[str | None, list[int]]:
    v0 = row[0] if row else None
    name = str(v0).strip() if v0 is not None else ""
    if not name:
        return None, []
    vals: list[int] = []
    for k in range(num_periods):
        cell = row[k + 1] if k + 1 < len(row) else None
        try:
            vals.append(int(cell) if cell is not None else 0)
        except (TypeError, ValueError):
            vals.append(0)
    return name, vals


def _parse_one_community_block(
    rows: list[tuple], start_i: int
) -> tuple[dict | None, int]:
    """start_i = offence title row; returns (block_dict, next_index)."""
    n = len(rows)
    if start_i + 1 >= n or not _is_quarter_header_row(rows[start_i + 1]):
        return None, start_i + 1
    category = str(rows[start_i][0]).strip()
    hdr = rows[start_i + 1]
    periods: list[str] = []
    for cell in hdr[1:]:
        if cell is None or (isinstance(cell, str) and not str(cell).strip()):
            break
        periods.append(str(cell).strip())
    if not periods:
        return None, start_i + 2
    num_p = len(periods)
    suburbs: list[dict] = []
    j = start_i + 2
    while j < n:
        r = rows[j]
        v0 = r[0] if r else None
        v0s = str(v0).strip() if v0 is not None else ""
        if not v0s:
            j += 1
            continue
        if v0s != "Total" and j + 1 < n and _offence_block_follows(rows, j):
            break
        name, vals = _parse_suburb_value_row(r, num_p)
        if not name:
            j += 1
            continue
        if len(vals) != num_p:
            while len(vals) < num_p:
                vals.append(0)
            vals = vals[:num_p]
        suburbs.append({"name": name, "q": vals})
        j += 1
        if name == "Total":
            break
    return {"category": category, "suburbs": suburbs}, j


def extract_community_quarterly(xlsx_path: Path) -> dict | None:
    """Parse Website_Qtrly_*.xlsx suburb-by-quarter blocks per policing district sheet."""
    if not xlsx_path.exists():
        return None
    wb = load_workbook(xlsx_path, read_only=True, data_only=True)
    promis = None
    districts_out: list[dict] = []

    for sheet_name in wb.sheetnames:
        district = COMMUNITY_SHEET_TO_DISTRICT.get(sheet_name)
        if not district:
            continue
        ws = wb[sheet_name]
        rows = [tuple(r) for r in ws.iter_rows(values_only=True)]
        if promis is None and len(rows) > 1:
            c0 = rows[1][0] if rows[1] else None
            if isinstance(c0, str) and "PROMIS" in c0:
                promis = c0.strip()
        periods_ref: list[str] | None = None
        blocks: list[dict] = []
        i = 0
        n = len(rows)
        while i < n - 1:
            a = rows[i][0] if rows[i] else None
            if not isinstance(a, str):
                i += 1
                continue
            title = a.strip()
            if not title or title == "Total":
                i += 1
                continue
            if "by Suburb" in title or title.startswith("PROMIS"):
                i += 1
                continue
            if not _offence_block_follows(rows, i):
                i += 1
                continue
            blk, nxt = _parse_one_community_block(rows, i)
            if blk and blk.get("suburbs"):
                if periods_ref is None:
                    periods_ref = []
                    h = rows[i + 1]
                    for cell in h[1:]:
                        if cell is None or (
                            isinstance(cell, str) and not str(cell).strip()
                        ):
                            break
                        periods_ref.append(str(cell).strip())
                blocks.append(
                    {
                        "category": blk["category"],
                        "suburbs": blk["suburbs"],
                    }
                )
            i = max(nxt, i + 1)
        if blocks and periods_ref:
            districts_out.append(
                {
                    "district": district,
                    "sourceSheet": sheet_name,
                    "categories": blocks,
                }
            )

    if not districts_out:
        return None
    periods_union = districts_out[0]["periodsChronological"]
    return {
        "sourceFile": xlsx_path.name,
        "granularity": "quarter",
        "promisAsAt": promis,
        "periodsChronological": periods_union,
        "districts": districts_out,
    }


def main() -> None:
    root = Path(__file__).resolve().parents[1]
    xlsx = root / "Website-Stats-Monthly-Mar26.xls.xlsx"
    out = root / "act-crime-viz" / "public" / "crime-data.json"
    if not xlsx.exists():
        print(f"Missing input: {xlsx}", file=sys.stderr)
        sys.exit(1)

    wb = load_workbook(xlsx, read_only=True, data_only=True)
    offence_ws = wb["Offence Statistics"]
    tables = extract_offence_tables(offence_ws)

    if not tables:
        print("No offence tables parsed", file=sys.stderr)
        sys.exit(1)

    traffic_ws = wb["Traffic Statistics - ACT"]
    traffic_title, traffic_periods, traffic_series = extract_simple_table(traffic_ws, 4)

    fv_ws = wb["Family Violence Statistics - AC"]
    fv_title, fv_periods, fv_series = extract_simple_table(fv_ws, 3)

    all_periods = merge_periods(
        *[t["periods"] for t in tables],
        traffic_periods,
        fv_periods,
    )

    sheets_meta = [
        {"sheetId": "offence", "sheetName": offence_ws.title, "tableCount": len(tables)},
        {"sheetId": "traffic", "sheetName": traffic_ws.title, "tableCount": 1},
        {"sheetId": "familyViolence", "sheetName": fv_ws.title, "tableCount": 1},
    ]

    tables_catalog: list[dict] = []
    for t in tables:
        tn = table_number_from_title(t.get("title"))
        tables_catalog.append(
            {
                "tableNumber": tn,
                "sheetId": "offence",
                "sheetName": offence_ws.title,
                "title": t.get("title"),
                "kind": "offence",
                "district": t["district"],
                "metricCount": len(t["series"]),
            }
        )
    tables_catalog.append(
        {
            "tableNumber": table_number_from_title(traffic_title),
            "sheetId": "traffic",
            "sheetName": traffic_ws.title,
            "title": traffic_title,
            "kind": "traffic",
            "district": "ACT",
            "metricCount": len(traffic_series),
        }
    )
    tables_catalog.append(
        {
            "tableNumber": table_number_from_title(fv_title),
            "sheetId": "familyViolence",
            "sheetName": fv_ws.title,
            "title": fv_title,
            "kind": "familyViolence",
            "district": "ACT",
            "metricCount": len(fv_series),
        }
    )

    for t in tables:
        t.pop("title", None)

    community_xlsx = root / "Website_Qtrly_Jun25.xlsx"
    community_quarterly = extract_community_quarterly(community_xlsx)

    payload = {
        "sourceFile": xlsx.name,
        "sheets": sheets_meta,
        "tablesCatalog": tables_catalog,
        "offenceTables": tables,
        "periodsChronological": all_periods,
        "traffic": {
            "tableNumber": 11,
            "title": traffic_title,
            "periods": traffic_periods,
            "series": traffic_series,
        },
        "familyViolence": {
            "tableNumber": 12,
            "title": fv_title,
            "periods": fv_periods,
            "series": fv_series,
        },
        "violenceOffenceKeys": [
            "Assault",
            "Homicide",
            "Sexual Assault",
            "Robbery",
            "Offences against a person",
        ],
        "communityQuarterly": community_quarterly,
    }

    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    print(f"Wrote {out}")


if __name__ == "__main__":
    main()
